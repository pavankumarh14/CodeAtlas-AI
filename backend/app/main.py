from fastapi import FastAPI, HTTPException, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import logging
import uuid
import datetime
import os
import zipfile
import io
import re
import requests

from .config import settings
from .graph import get_graph_driver
from .vectorstore import get_vector_store
from .agents import AgentOrchestrator
from .data import seed_all_data
from .repository_ingestion import inspect_public_repository, MANIFEST_NAMES, DOC_SUFFIXES

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title=settings.PROJECT_NAME,
    description="The Living Engineering Ontology and Knowledge Graph Platform API"
)

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory query activity log
activity_log: List[Dict[str, Any]] = []

# Instantiate Orchestrator
orchestrator = AgentOrchestrator()

class AnalyzeRequest(BaseModel):
    query: str

class RepositoryImportRequest(BaseModel):
    repository_url: str

@app.on_event("startup")
def startup_event():
    # Attempt to establish connections and auto-seed if graph database is empty
    try:
        driver = get_graph_driver()
        nodes = driver.get_nodes()
        if not nodes:
            logger.info("Graph database is empty. Auto-seeding initial data...")
            seed_all_data()
    except Exception as e:
        logger.error(f"Startup database connection or seed failed: {e}")

@app.post("/api/v1/seed", tags=["Data"])
def trigger_seed():
    try:
        seed_all_data()
        return {"status": "success", "message": "Database seeded successfully."}
    except Exception as e:
        logger.error(f"Manual seed trigger failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/stats", tags=["Dashboard"])
def get_dashboard_stats():
    try:
        driver = get_graph_driver()
        services = len(driver.get_nodes("Service"))
        repos = len(driver.get_nodes("Repository"))
        teams = len(driver.get_nodes("Team"))
        engineers = len(driver.get_nodes("Engineer"))
        requirements = len(driver.get_nodes("Requirement"))
        
        all_incidents = driver.get_nodes("Incident")
        active_incidents = sum(1 for inc in all_incidents if inc["properties"].get("status") == "Active")
        
        # Calculate system health score
        # Base health is 100%, deduct points for active incidents and missing compliance
        health_score = max(30, 100 - (active_incidents * 15))
        
        return {
            "services": services,
            "repositories": repos,
            "teams": teams,
            "engineers": engineers,
            "requirements": requirements,
            "active_incidents": active_incidents,
            "system_health_score": health_score
        }
    except Exception as e:
        logger.error(f"Failed to get dashboard stats: {e}")
        return {
            "services": 0, "repositories": 0, "teams": 0, "engineers": 0,
            "requirements": 0, "active_incidents": 0, "system_health_score": 100
        }

@app.get("/api/v1/incidents", tags=["Dashboard"])
def list_incidents():
    try:
        driver = get_graph_driver()
        incidents = driver.get_nodes("Incident")
        return [inc["properties"] for inc in incidents]
    except Exception as e:
        logger.error(f"Failed to list incidents: {e}")
        return []

@app.get("/api/v1/requirements", tags=["Dashboard"])
def list_requirements():
    try:
        driver = get_graph_driver()
        reqs = driver.get_nodes("Requirement")
        return [r["properties"] for r in reqs]
    except Exception as e:
        logger.error(f"Failed to list requirements: {e}")
        return []

@app.post("/api/v1/analyze", tags=["Agents"])
def analyze_query(request: AnalyzeRequest):
    if not request.query.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty.")
    
    start_time = datetime.datetime.now()
    try:
        result = orchestrator.execute(request.query)
        end_time = datetime.datetime.now()
        duration_ms = int((end_time - start_time).total_seconds() * 1000)
        
        log_entry = {
            "id": str(uuid.uuid4()),
            "timestamp": start_time.isoformat(),
            "query": request.query,
            "flow_type": result.get("flow_type"),
            "target_agent": result.get("target_agent"),
            "duration_ms": duration_ms,
            "status": "Success"
        }
        activity_log.insert(0, log_entry)
        
        return result
    except Exception as e:
        logger.error(f"Error in orchestrator execution: {e}", exc_info=True)
        log_entry = {
            "id": str(uuid.uuid4()),
            "timestamp": start_time.isoformat(),
            "query": request.query,
            "flow_type": "unknown",
            "target_agent": "Orchestrator",
            "duration_ms": 0,
            "status": f"Error: {str(e)}"
        }
        activity_log.insert(0, log_entry)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/repositories/import", tags=["Repository Intake"])
def import_public_repository(request: RepositoryImportRequest):
    """Inspect public GitHub metadata and add the discoverable evidence to the graph."""
    try:
        profile = inspect_public_repository(request.repository_url)
        graph = get_graph_driver()
        graph.add_node("Repository", {
            "name": profile["repository"],
            "url": profile["url"],
            "description": profile["description"],
            "default_branch": profile["default_branch"],
            "languages": ", ".join(profile["languages"]),
            "source": "Public GitHub import",
        })
        for path in profile["documents"]:
            title = f"{profile['repository']} · {path}"
            graph.add_node("Document", {"title": title, "path": path, "repository": profile["repository"], "source": "Public GitHub import"})
            graph.add_relationship("Repository", profile["repository"], "Document", title, "CONTAINS")

        activity_log.insert(0, {
            "id": str(uuid.uuid4()), "timestamp": datetime.datetime.now().isoformat(),
            "query": f"Imported public repository {profile['repository']}", "flow_type": "repository_intake",
            "target_agent": "Repository Intake", "duration_ms": 0, "status": "Success"
        })
        return {**profile, "graph_nodes_added": 1 + len(profile["documents"])}
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error))
    except requests.RequestException:
        logger.exception("Public repository lookup failed")
        raise HTTPException(status_code=502, detail="Could not reach GitHub. Please try again.")

DOC_SUFFIXES = (".md", ".mdx", ".rst", ".txt", ".csv", ".xlsx", ".xls")

def extract_file_content(filename: str, file_bytes: bytes) -> Optional[str]:
    lower_name = filename.lower()
    if lower_name.endswith((".md", ".mdx", ".rst", ".txt")):
        return file_bytes.decode("utf-8", errors="ignore")
    elif lower_name.endswith(".csv"):
        try:
            import csv
            content = file_bytes.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(content))
            text_content = []
            for row in reader:
                row_str = ", ".join([str(val) if val is not None else "" for val in row])
                if row_str.strip(", "):
                    text_content.append(row_str)
            return "\n".join(text_content)
        except Exception as e:
            logger.error(f"Error parsing CSV {filename}: {e}")
            return None
    elif lower_name.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            text_content = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_content.append(f"Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_str = ", ".join([str(val) if val is not None else "" for val in row])
                    if row_str.strip(", "):
                        text_content.append(row_str)
            return "\n".join(text_content)
        except Exception as e:
            logger.error(f"Error parsing Excel {filename}: {e}")
            return None
    return None

def extract_tables_from_zip(body: bytes) -> List[tuple[str, List[str], List[List[Any]]]]:
    tables = []
    with zipfile.ZipFile(io.BytesIO(body)) as z:
        for path in z.namelist():
            if path.endswith("/") or "__MACOSX" in path:
                continue
            lower_name = path.lower()
            if lower_name.endswith((".xlsx", ".xls")):
                try:
                    import openpyxl
                    with z.open(path) as f:
                        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            rows = []
                            for row in sheet.iter_rows(values_only=True):
                                rows.append(list(row))
                            if len(rows) > 1:
                                headers = [str(h).lower().strip() if h is not None else "" for h in rows[0]]
                                data_rows = rows[1:]
                                tables.append((f"{path}:{sheet_name}", headers, data_rows))
                except Exception as e:
                    logger.error(f"Error reading excel table {path}: {e}")
            elif lower_name.endswith(".csv"):
                try:
                    import csv
                    with z.open(path) as f:
                        content = f.read().decode("utf-8", errors="ignore")
                        reader = csv.reader(io.StringIO(content))
                        rows = list(reader)
                        if len(rows) > 1:
                            headers = [str(h).lower().strip() if h is not None else "" for h in rows[0]]
                            data_rows = rows[1:]
                            tables.append((path, headers, data_rows))
                except Exception as e:
                    logger.error(f"Error reading CSV table {path}: {e}")
    return tables

def parse_table_and_add_to_graph(table_name: str, headers: List[str], rows: List[List[Any]], graph):
    table_name = table_name.lower().strip()
    
    # 1. TEAM
    if "team" in table_name:
        name_idx = next((i for i, h in enumerate(headers) if "name" in h or "team" in h), -1)
        if name_idx != -1:
            for r in rows:
                if len(r) > name_idx and r[name_idx]:
                    team_name = str(r[name_idx]).strip()
                    props = {headers[i]: str(val).strip() for i, val in enumerate(r) if val is not None and i < len(headers)}
                    props["name"] = team_name
                    graph.add_node("Team", props)
                    
    # 2. ENGINEER
    elif "engineer" in table_name or "member" in table_name or "people" in table_name or "person" in table_name:
        name_idx = next((i for i, h in enumerate(headers) if "name" in h or "engineer" in h), -1)
        team_idx = next((i for i, h in enumerate(headers) if "team" in h), -1)
        if name_idx != -1:
            for r in rows:
                if len(r) > name_idx and r[name_idx]:
                    eng_name = str(r[name_idx]).strip()
                    props = {headers[i]: str(val).strip() for i, val in enumerate(r) if val is not None and i < len(headers)}
                    props["name"] = eng_name
                    graph.add_node("Engineer", props)
                    if team_idx != -1 and len(r) > team_idx and r[team_idx]:
                        team_name = str(r[team_idx]).strip()
                        graph.add_relationship("Engineer", eng_name, "Team", team_name, "MEMBER_OF")

    # 3. SERVICE
    elif "service" in table_name:
        name_idx = next((i for i, h in enumerate(headers) if "name" in h or "service" in h), -1)
        owner_idx = next((i for i, h in enumerate(headers) if "owner" in h or "team" in h), -1)
        repo_idx = next((i for i, h in enumerate(headers) if "repo" in h or "repository" in h), -1)
        if name_idx != -1:
            for r in rows:
                if len(r) > name_idx and r[name_idx]:
                    svc_name = str(r[name_idx]).strip()
                    props = {headers[i]: str(val).strip() for i, val in enumerate(r) if val is not None and i < len(headers)}
                    props["name"] = svc_name
                    graph.add_node("Service", props)
                    if owner_idx != -1 and len(r) > owner_idx and r[owner_idx]:
                        team_name = str(r[owner_idx]).strip()
                        graph.add_relationship("Team", team_name, "Service", svc_name, "OWNS")
                    if repo_idx != -1 and len(r) > repo_idx and r[repo_idx]:
                        repo_name = str(r[repo_idx]).strip()
                        graph.add_relationship("Repository", repo_name, "Service", svc_name, "CONTAINS")

    # 4. REPOSITORY
    elif "repo" in table_name:
        name_idx = next((i for i, h in enumerate(headers) if "name" in h or "repo" in h), -1)
        if name_idx != -1:
            for r in rows:
                if len(r) > name_idx and r[name_idx]:
                    repo_name = str(r[name_idx]).strip()
                    props = {headers[i]: str(val).strip() for i, val in enumerate(r) if val is not None and i < len(headers)}
                    props["name"] = repo_name
                    graph.add_node("Repository", props)

    # 5. INCIDENT
    elif "incident" in table_name or "bug" in table_name or "outage" in table_name:
        id_idx = next((i for i, h in enumerate(headers) if "id" in h or "incident" in h or "bug" in h), -1)
        title_idx = next((i for i, h in enumerate(headers) if "title" in h or "name" in h or "summary" in h), -1)
        svc_idx = next((i for i, h in enumerate(headers) if "service" in h or "app" in h), -1)
        if id_idx != -1:
            for r in rows:
                if len(r) > id_idx and r[id_idx]:
                    inc_id = str(r[id_idx]).strip()
                    title = str(r[title_idx]).strip() if (title_idx != -1 and len(r) > title_idx and r[title_idx]) else f"Incident {inc_id}"
                    props = {headers[i]: str(val).strip() for i, val in enumerate(r) if val is not None and i < len(headers)}
                    props["inc_id"] = inc_id
                    props["title"] = title
                    graph.add_node("Incident", props)
                    if svc_idx != -1 and len(r) > svc_idx and r[svc_idx]:
                        svc_name = str(r[svc_idx]).strip()
                        graph.add_relationship("Incident", inc_id, "Service", svc_name, "AFFECTS")

    # 6. DEPENDENCY / RELATIONSHIP
    elif "dependency" in table_name or "dependencies" in table_name or "relation" in table_name or "relationship" in table_name:
        src_idx = next((i for i, h in enumerate(headers) if "source" in h or "from" in h or "service" in h), -1)
        tgt_idx = next((i for i, h in enumerate(headers) if "target" in h or "to" in h or "depends" in h or "uses" in h), -1)
        type_idx = next((i for i, h in enumerate(headers) if "type" in h or "relation" in h), -1)
        if src_idx != -1 and tgt_idx != -1:
            for r in rows:
                if len(r) > src_idx and r[src_idx] and len(r) > tgt_idx and r[tgt_idx]:
                    src_name = str(r[src_idx]).strip()
                    tgt_name = str(r[tgt_idx]).strip()
                    rel_type = str(r[type_idx]).strip().upper() if (type_idx != -1 and len(r) > type_idx and r[type_idx]) else "DEPENDS_ON"
                    
                    src_label = "Service"
                    tgt_label = "Service"
                    if ":" in src_name:
                        src_label, src_name = src_name.split(":", 1)
                    if ":" in tgt_name:
                        tgt_label, tgt_name = tgt_name.split(":", 1)
                        
                    graph.add_relationship(src_label.strip(), src_name.strip(), tgt_label.strip(), tgt_name.strip(), rel_type)

def parse_generic_table_to_graph(repo_name: str, table_name: str, headers: List[str], rows: List[List[Any]], graph) -> int:
    display_name = table_name.rsplit("/", 1)[-1]
    
    # 1. Create a Table node
    graph.add_node("Table", {
        "name": display_name,
        "type": "Spreadsheet Table",
        "columns": ", ".join(headers),
        "rows_count": len(rows)
    })
    graph.add_relationship("Repository", repo_name, "Table", display_name, "CONTAINS")
    
    nodes_added = 1
    
    # 2. Add each row as a Row node
    for idx, r in enumerate(rows):
        row_props = {}
        for col_idx, val in enumerate(r):
            if val is not None and col_idx < len(headers):
                col_name = headers[col_idx]
                if col_name:
                    row_props[col_name] = str(val).strip()
                    
        if not row_props:
            continue
            
        id_keys = ["id", "name", "title", "workflow", "feature", "key", "number", "requirement"]
        row_name = ""
        for key in id_keys:
            match_col = next((k for k in row_props.keys() if key in k.lower()), None)
            if match_col:
                row_name = row_props[match_col]
                break
                
        if not row_name:
            row_name = f"Row {idx + 1}"
            
        row_node_id = f"{display_name} · Row {idx + 1}"
        row_props["name"] = row_name
        row_props["id"] = row_node_id
        row_props["table"] = display_name
        
        # Add node
        graph.add_node("Row", row_props)
        nodes_added += 1
        
        # Link Table -> HAS_ROW -> Row
        graph.add_relationship("Table", display_name, "Row", row_name, "HAS_ROW")
        
        # Handle explicit relations if any columns match relations
        rel_keys = ["depends", "parent", "related", "links", "uses", "requires", "next"]
        for key in rel_keys:
            match_col = next((k for k in row_props.keys() if key in k.lower()), None)
            if match_col:
                target_val = row_props[match_col]
                if target_val and str(target_val).lower() not in ("none", "n/a", "", "null"):
                    graph.add_relationship("Row", row_name, "Row", str(target_val).strip(), "LINKS_TO")
                    
    return nodes_added

@app.post("/api/v1/repositories/upload", tags=["Repository Intake"])
async def upload_repository_zip(request: Request, filename: str):
    """Stage and inspect a local repository uploaded as a ZIP archive, adding it to the graph."""
    try:
        body = await request.body()
        if not body:
            raise HTTPException(status_code=400, detail="Empty file uploaded.")
            
        repo_name = filename
        if repo_name.lower().endswith(".zip"):
            repo_name = repo_name[:-4]
            
        # Read zip in-memory
        try:
            with zipfile.ZipFile(io.BytesIO(body)) as z:
                paths = [info.filename for info in z.infolist() if not info.is_dir()][:1000]
        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="Invalid ZIP file structure.")
            
        # Classify files
        manifests = [path for path in paths if path.rsplit("/", 1)[-1].lower() in MANIFEST_NAMES][:12]
        documents = [path for path in paths if path.lower().endswith(DOC_SUFFIXES)][:50]
        api_candidates = [path for path in paths if re.search(r"(^|/)(api|routes?|controllers?)(/|$)|/(route|controller)\.", path, re.IGNORECASE)][:12]
        
        # Detect languages based on file extensions
        lang_map = {
            ".py": "Python",
            ".js": "JavaScript",
            ".ts": "TypeScript",
            ".tsx": "TypeScript",
            ".jsx": "JavaScript",
            ".go": "Go",
            ".java": "Java",
            ".rb": "Ruby",
            ".rs": "Rust",
            ".cs": "C#",
            ".cpp": "C++",
            ".c": "C",
            ".xlsx": "Excel",
            ".xls": "Excel",
            ".csv": "CSV",
        }
        detected_langs = set()
        for path in paths:
            ext = "." + path.rsplit(".", 1)[-1].lower() if "." in path else ""
            if ext in lang_map:
                detected_langs.add(lang_map[ext])
                
        languages = list(detected_langs)[:8]
        
        profile = {
            "repository": repo_name,
            "url": f"local://{filename}",
            "description": f"Local repository staged via ZIP upload of {filename}.",
            "default_branch": "main",
            "languages": languages,
            "files_scanned": len(paths),
            "tree_truncated": len(paths) >= 1000,
            "manifests": manifests,
            "documents": documents,
            "api_candidates": api_candidates,
            "graph_nodes_added": 1,
        }
        
        # Add to graph
        graph = get_graph_driver()
        graph.add_node("Repository", {
            "name": profile["repository"],
            "url": profile["url"],
            "description": profile["description"],
            "default_branch": profile["default_branch"],
            "languages": ", ".join(profile["languages"]),
            "source": "Local ZIP upload",
        })
        
        vector_store = get_vector_store()
        nodes_added = 1
        
        # Single-pass: open zip once, read all files, process content + tables together
        try:
            with zipfile.ZipFile(io.BytesIO(body)) as z:
                all_names = [info.filename for info in z.infolist() if not info.is_dir() and "__MACOSX" not in info.filename]
                
                # Track tables extracted from Excel/CSV for graph insertion
                parsed_tables: List[tuple] = []
                
                for path in all_names:
                    lower_path = path.lower()
                    is_text_doc = lower_path.endswith((".md", ".mdx", ".rst", ".txt"))
                    is_excel = lower_path.endswith((".xlsx", ".xls"))
                    is_csv = lower_path.endswith(".csv")
                    
                    if not (is_text_doc or is_excel or is_csv):
                        continue

                    try:
                        with z.open(path) as f:
                            file_bytes = f.read()
                    except Exception as ex:
                        logger.error(f"Could not read {path} from ZIP: {ex}")
                        continue

                    # ── TEXT / MARKDOWN docs ─────────────────────────────────
                    if is_text_doc:
                        content = file_bytes.decode("utf-8", errors="ignore")
                        if content.strip():
                            vector_store.add_texts(
                                texts=[content],
                                metadatas=[{"source": f"{repo_name} · {path}", "repository": repo_name, "path": path, "origin": "uploaded"}],
                                ids=[f"doc:{repo_name}:{path}"]
                            )
                            title = f"{repo_name} · {path}"
                            graph.add_node("Document", {
                                "title": title, "path": path,
                                "repository": repo_name, "source": "Local ZIP upload",
                                "content_summary": content[:500]
                            })
                            graph.add_relationship("Repository", repo_name, "Document", title, "CONTAINS")
                            nodes_added += 1

                    # ── EXCEL sheets ─────────────────────────────────────────
                    elif is_excel:
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                            for sheet_name in wb.sheetnames:
                                sheet = wb[sheet_name]
                                all_rows = []
                                for row in sheet.iter_rows(values_only=True):
                                    if any(cell is not None for cell in row):
                                        all_rows.append(list(row))
                                
                                if len(all_rows) < 1:
                                    continue

                                # Convert entire sheet to text for vector indexing
                                text_lines = [f"File: {path}, Sheet: {sheet_name}"]
                                for row in all_rows:
                                    line = ", ".join([str(v) if v is not None else "" for v in row])
                                    if line.strip(", "):
                                        text_lines.append(line)
                                sheet_text = "\n".join(text_lines)

                                sheet_id = f"{path}:{sheet_name}"
                                vector_store.add_texts(
                                    texts=[sheet_text],
                                    metadatas=[{"source": sheet_id, "repository": repo_name, "path": path, "sheet": sheet_name, "origin": "uploaded"}],
                                    ids=[f"doc:{repo_name}:{sheet_id}"]
                                )
                                logger.info(f"Indexed sheet '{sheet_name}' from '{path}' into vector store ({len(all_rows)} rows)")

                                # Add Document node for the sheet
                                title = f"{repo_name} · {sheet_id}"
                                graph.add_node("Document", {
                                    "title": title, "path": path, "sheet": sheet_name,
                                    "repository": repo_name, "source": "Local ZIP upload",
                                    "content_summary": sheet_text[:500]
                                })
                                graph.add_relationship("Repository", repo_name, "Document", title, "CONTAINS")
                                nodes_added += 1

                                # Collect for table-to-graph parsing
                                if len(all_rows) > 1:
                                    headers = [str(h).lower().strip() if h is not None else f"col{i}" for i, h in enumerate(all_rows[0])]
                                    parsed_tables.append((sheet_id, headers, all_rows[1:]))
                        except Exception as ex:
                            logger.error(f"Error parsing Excel '{path}': {ex}")

                    # ── CSV ──────────────────────────────────────────────────
                    elif is_csv:
                        try:
                            import csv
                            content_str = file_bytes.decode("utf-8", errors="ignore")
                            rows = list(csv.reader(io.StringIO(content_str)))
                            if rows:
                                sheet_text = f"File: {path}\n" + "\n".join([", ".join(r) for r in rows if any(c.strip() for c in r)])
                                vector_store.add_texts(
                                    texts=[sheet_text],
                                    metadatas=[{"source": path, "repository": repo_name, "path": path, "origin": "uploaded"}],
                                    ids=[f"doc:{repo_name}:{path}"]
                                )
                                logger.info(f"Indexed CSV '{path}' into vector store ({len(rows)} rows)")
                                title = f"{repo_name} · {path}"
                                graph.add_node("Document", {
                                    "title": title, "path": path,
                                    "repository": repo_name, "source": "Local ZIP upload",
                                    "content_summary": sheet_text[:500]
                                })
                                graph.add_relationship("Repository", repo_name, "Document", title, "CONTAINS")
                                nodes_added += 1
                                if len(rows) > 1:
                                    headers = [str(h).lower().strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
                                    parsed_tables.append((path, headers, rows[1:]))
                        except Exception as ex:
                            logger.error(f"Error parsing CSV '{path}': {ex}")

                # ── Graph structural + generic table parsers ──────────────
                for table_name, headers, data_rows in parsed_tables:
                    try:
                        parse_table_and_add_to_graph(table_name, headers, data_rows, graph)
                    except Exception as ex:
                        logger.error(f"Ontological parse failed for '{table_name}': {ex}")
                    try:
                        tbl_nodes = parse_generic_table_to_graph(repo_name, table_name, headers, data_rows, graph)
                        nodes_added += tbl_nodes
                    except Exception as ex:
                        logger.error(f"Generic table parse failed for '{table_name}': {ex}")

        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="Invalid ZIP file structure.")
        except Exception as ex:
            logger.exception("Error during ZIP content processing")
            raise HTTPException(status_code=500, detail=str(ex))

        profile["graph_nodes_added"] = nodes_added
        
        # Add to activity log
        activity_log.insert(0, {
            "id": str(uuid.uuid4()), "timestamp": datetime.datetime.now().isoformat(),
            "query": f"Uploaded local repository {profile['repository']} and parsed structure", "flow_type": "repository_intake",
            "target_agent": "Repository Intake", "duration_ms": 0, "status": "Success"
        })
        
        return profile
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("ZIP upload failed")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/database/clear", tags=["Data"])
def clear_database():
    """Clear all nodes, relationships, and vector index entries to start fresh."""
    try:
        driver = get_graph_driver()
        driver.clear()
        
        vector_store = get_vector_store()
        vector_store.clear()
        
        activity_log.clear()
        
        return {"status": "success", "message": "Database and Vector Store cleared successfully."}
    except Exception as e:
        logger.error(f"Failed to clear database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/agent-activity-log", tags=["Agents"])
def get_activity_log():
    return activity_log

@app.get("/api/v1/knowledge-gaps", tags=["Agents"])
def scan_knowledge_gaps():
    try:
        agent = orchestrator.agents["knowledge_gap"]
        res = agent.run("Scan for knowledge gaps")
        return res["result"]
    except Exception as e:
        logger.error(f"Failed to scan knowledge gaps: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/experts", tags=["Agents"])
def search_experts(service: str):
    try:
        agent = orchestrator.agents["expert_discovery"]
        res = agent.run(f"Who owns {service}")
        return res["result"]
    except Exception as e:
        logger.error(f"Failed to search experts: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/v1/graph/data", tags=["Graph"])
def get_graph_data():
    """Retrieve nodes and edges formatted for React Flow rendering."""
    try:
        driver = get_graph_driver()
        raw_nodes = driver.get_nodes()
        raw_rels = driver.get_relationships()
        
        formatted_nodes = []
        formatted_edges = []
        
        # Color mapping for UI ontology nodes
        # React Flow custom node styles
        color_map = {
            "Team": {"background": "#FFF3CD", "border": "#FFC107", "color": "#856404"},       # Yellow
            "Engineer": {"background": "#D4EDDA", "border": "#28A745", "color": "#155724"},   # Green
            "Service": {"background": "#CCE5FF", "border": "#007BFF", "color": "#004085"},    # Blue
            "Repository": {"background": "#E8D9F2", "border": "#9C27B0", "color": "#4A0072"}, # Purple
            "API": {"background": "#D1ECF1", "border": "#17A2B8", "color": "#0C5460"},        # Teal
            "Requirement": {"background": "#F8D7DA", "border": "#DC3545", "color": "#721C24"},# Red/Pink
            "Incident": {"background": "#FCE8E6", "border": "#EA4335", "color": "#C5221F"},   # Orange/Red
            "Document": {"background": "#E2E3E5", "border": "#6C757D", "color": "#383D41"},   # Grey
            "Runbook": {"background": "#E2F0D9", "border": "#70AD47", "color": "#385623"}     # Light Green
        }
        
        # Grid layout helper coordinates to prevent overlapping
        category_x = {
            "Team": 100,
            "Engineer": 100,
            "Service": 400,
            "Repository": 700,
            "API": 700,
            "Requirement": 400,
            "Incident": 400,
            "Document": 950,
            "Runbook": 950
        }
        
        y_counters = {k: 50 for k in category_x.keys()}
        
        for node in raw_nodes:
            lbl = node["labels"][0]
            properties = node["properties"]
            
            # Primary naming extraction
            name = properties.get("name") or properties.get("title") or properties.get("req_id") or properties.get("inc_id") or "Unnamed Node"
            
            # Position allocation
            x = category_x.get(lbl, 500)
            if lbl == "Engineer":
                x = 100
                y = y_counters["Team"] + 50
                y_counters["Team"] = y  # Stagger them
            else:
                y = y_counters.get(lbl, 100)
                y_counters[lbl] = y + 120
            
            style = color_map.get(lbl, {"background": "#FFFFFF", "border": "#CCCCCC", "color": "#333333"})
            
            formatted_nodes.append({
                "id": node["id"],
                "type": "customNode",  # React Flow custom node
                "position": {"x": x, "y": y},
                "data": {
                    "label": name,
                    "type": lbl,
                    "properties": properties,
                    "style": style
                }
            })
            
        for rel in raw_rels:
            formatted_edges.append({
                "id": f"edge-{rel['id']}",
                "source": rel["start_node"]["id"],
                "target": rel["end_node"]["id"],
                "label": rel["type"],
                "animated": rel["type"] in ("DEPENDS_ON", "USES"),
                "style": {"stroke": "#999999", "strokeWidth": 2}
            })
            
        return {
            "nodes": formatted_nodes,
            "edges": formatted_edges
        }
    except Exception as e:
        logger.error(f"Failed to fetch graph data: {e}", exc_info=True)
        return {"nodes": [], "edges": []}

# In the Render image, the Next.js static export is copied here during the Docker build.
# Mount this last so all API and documentation routes remain owned by FastAPI.
static_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static")
if os.path.isdir(static_dir):
    app.mount("/", StaticFiles(directory=static_dir, html=True), name="frontend")
